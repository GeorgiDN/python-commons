from openpyxl import *
from datetime import datetime, timedelta

# Create a new workbook
wb = load_workbook('Register.xlsx')
sheet = wb.active


def create_new_sheet(sheet_name=None):

    if sheet_name is None:
        return wb.create_sheet()

    return wb.create_sheet(title=sheet_name)


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Time"
    sheet.cell(row=1, column=2).value = "First Name"
    sheet.cell(row=1, column=3).value = "Last Name"
    sheet.cell(row=1, column=4).value = "Phone Number"
    sheet.cell(row=1, column=5).value = "Address"
    sheet.cell(row=1, column=6).value = "Email"
    sheet.cell(row=1, column=7).value = "Password"


def add_data_to_excel(time_, first_name, last_name, phone_number, address, email, password):

    # Find the next empty row in the Excel sheet
    next_row = sheet.max_row + 1

    # Write food data to the Excel sheet
    sheet.cell(row=next_row, column=1).value = time_
    sheet.cell(row=next_row, column=2).value = first_name
    sheet.cell(row=next_row, column=3).value = last_name
    sheet.cell(row=next_row, column=4).value = phone_number
    sheet.cell(row=next_row, column=5).value = address
    sheet.cell(row=next_row, column=6).value = email
    sheet.cell(row=next_row, column=7).value = password

    wb.save('Register.xlsx')


excel()
