import os
import openpyxl

# Directory containing the images
image_folder = 'main_folder'

# Create a new Excel workbook and select the active sheet
wb = openpyxl.Workbook()
sheet = wb.active

# Set the headers in the Excel file
sheet['A1'] = 'Number'
sheet['B1'] = 'Link to Image'

# Iterate over the files in the folder
row = 2  # Start writing from the second row (after the headers)
for filename in os.listdir(image_folder):
    if filename.endswith(('.jpg', '.png')):  # Adjust file types as needed
        # Extract the number after "T_"
        number = filename.split('_')[1][:-4]

        # Write the number to column A
        sheet[f'A{row}'] = number

        # Create the clickable link to the image in column B
        file_path = os.path.join(image_folder, filename)
        # link = f"=HYPERLINK(\"{file_path}\", \"{filename}\")"
        link = f"=HYPERLINK(\"{filename}\")"
        sheet[f'B{row}'] = link

        row += 1

# Save the Excel file in the "main_folder" folder
excel_file = os.path.join(image_folder, 'output_with_links.xlsx')
wb.save(excel_file)

print(f"Data and links have been written to {excel_file}")





# # Without link
# import os
# import openpyxl
#
# image_folder = 'main_folder'
#
# wb = openpyxl.Workbook()
# sheet = wb.active
#
# sheet['A1'] = 'Number'
# sheet['B1'] = 'Filename'
#
#
# row = 2  # Start writing from the second row (after the headers)
# for filename in os.listdir(image_folder):
#     if filename.endswith(('.jpg', '.png')):  # Adjust file types as needed
#         # Extract the number after "T_"
#         number = filename.split('_')[1][:-4]
#
#         # Write the number and the filename to the Excel file
#         sheet[f'A{row}'] = number
#         sheet[f'B{row}'] = filename
#         row += 1
#
# # Save the Excel file
# excel_file = 'output.xlsx'
# wb.save(excel_file)
#
# print(f"Data has been written to {excel_file}")





# # ADD picture in the file
# import os
# from PIL import Image as PilImage  # Import from Pillow
# import openpyxl
# from openpyxl.drawing.image import Image as ExcelImage
#
# # Directory containing the images
# image_folder = 'main_folder'
#
# # Create a new Excel workbook and select the active sheet
# wb = openpyxl.Workbook()
# sheet = wb.active
#
# # Set the headers in the Excel file
# sheet['A1'] = 'Number'
# sheet['B1'] = 'Filename'
# sheet['C1'] = 'Image'
#
# # Resize factor for images
# resize_factor = 0.1  # Resize to 50% of the original size
#
# # Iterate over the files in the folder
# row = 2  # Start writing from the second row (after the headers)
# for filename in os.listdir(image_folder):
#     if filename.endswith(('.jpg', '.png')):  # Adjust file types as needed
#         # Extract the number after "T_"
#         number = filename.split('_')[1][:-4]
#
#         # Write the number and the filename to the Excel file
#         sheet[f'A{row}'] = number
#         sheet[f'B{row}'] = filename
#
#         # Full path to the image
#         image_path = os.path.join(image_folder, filename)
#
#         # Open the image using Pillow
#         img = PilImage.open(image_path)
#
#         # Resize the image
#         img = img.resize((int(img.width * resize_factor), int(img.height * resize_factor)))
#
#         # Save the resized image temporarily
#         resized_image_path = f'resized_{filename}'
#         img.save(resized_image_path)
#
#         # Add the resized image to the Excel file
#         excel_img = ExcelImage(resized_image_path)
#         sheet.add_image(excel_img, f'C{row}')
#
#         # Dynamically adjust the row height based on the resized image height (convert pixels to points)
#         sheet.row_dimensions[row].height = excel_img.height * 0.75
#
#         # Dynamically adjust the column width based on the resized image width (convert pixels to Excel units)
#         sheet.column_dimensions['C'].width = excel_img.width * 0.14
#
#         row += 1
#
# # Save the Excel file
# excel_file = 'output_with_resized_images.xlsx'
# wb.save(excel_file)
#
# print(f"Data and resized images have been written to {excel_file}")

