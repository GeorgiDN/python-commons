import os
import openpyxl

# Directory containing the subfolders and images
image_folder = 'main_folder'

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Iterate over the subfolders in the image folder
for foldername in os.listdir(image_folder):
    folder_path = os.path.join(image_folder, foldername)

    # Check if it's a folder
    if os.path.isdir(folder_path):
        # Create a new sheet with the name of the folder
        sheet = wb.create_sheet(title=foldername)

        # Set the headers in the Excel sheet
        sheet['A1'] = 'Description'
        sheet['B1'] = 'Link to Image'

        # Iterate over the files in the folder
        row = 2  # Start writing from the second row (after the headers)
        for filename in os.listdir(folder_path):
            if filename.endswith(('.jpg', '.png')):  # Adjust file types as needed
                number = filename.split('_')[1]

                # Write the number to column A
                sheet[f'A{row}'] = number

                # Create the clickable link to the image in column B
                file_path = os.path.join(folder_path, filename)  # Full path to the image
                link = f"=HYPERLINK(\"{file_path}\", \"{filename[:-4]}\")"
                # link = f"=HYPERLINK(\"{filename}\")"
                sheet[f'B{row}'] = link

                row += 1

# Remove the default sheet created by openpyxl if it's empty
if 'Sheet' in wb.sheetnames and wb['Sheet'].max_row == 1:
    wb.remove(wb['Sheet'])

# Save the Excel file in the "main_folder" folder
excel_file = os.path.join(image_folder, 'output.xlsx')
wb.save(excel_file)

print(f"Data and links have been written to {excel_file}")
