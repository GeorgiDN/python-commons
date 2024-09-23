from openpyxl import *
from datetime import datetime, timedelta

# Create a new workbook
wb = load_workbook('Food_data.xlsx')
sheet = wb.active


def create_new_sheet(sheet_name=None):

    if sheet_name is None:
        return wb.create_sheet()

    return wb.create_sheet(title=sheet_name)


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 10

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Time"
    sheet.cell(row=1, column=2).value = "Food Name"
    sheet.cell(row=1, column=3).value = "Calories"
    sheet.cell(row=1, column=4).value = "Protein"
    sheet.cell(row=1, column=5).value = "Carbs"
    sheet.cell(row=1, column=6).value = "Fats"


def add_food_to_excel(time, food_name, calories, protein, carbs, fats):

    # Find the next empty row in the Excel sheet
    next_row = sheet.max_row + 1

    # Write food data to the Excel sheet
    sheet.cell(row=next_row, column=1).value = time
    sheet.cell(row=next_row, column=2).value = food_name
    sheet.cell(row=next_row, column=3).value = calories
    sheet.cell(row=next_row, column=4).value = protein
    sheet.cell(row=next_row, column=5).value = carbs
    sheet.cell(row=next_row, column=6).value = fats

    wb.save('Food_data.xlsx')


excel()
